"""
AI Exam Manager — Excel Import/Export Handler
Handles .xlsx file reading and writing for all data sections.
"""
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO


# Column definitions for each section
SECTION_COLUMNS = {
    'invigilators': ['Name', 'Email', 'Phone', 'Department', 'Available (Yes/No)', 'Max Duties', 'Group ID'],
    'rooms': ['Name', 'Capacity', 'Floor', 'Room Type', 'Available (Yes/No)', 'Equipment', 'Buffer Seats'],
    'subjects': ['Name', 'Code', 'Branch', 'Student Count', 'Color'],
    'students': ['Name', 'Roll No', 'Branch', 'Group ID', 'Email', 'Phone'],
    'inventory': ['Name', 'Category', 'Quantity', 'Min Threshold', 'Unit'],
    'branches': ['Name', 'Code', 'Color', 'Student Count'],
    'staff_duties': ['Staff Name', 'Duty Description', 'Location', 'Date', 'Start Time', 'End Time'],
    'exams': ['Subject', 'Room', 'Date', 'Start Time', 'End Time', 'Session', 'Status'],
}


def _style_header(ws, num_cols):
    """Apply styling to the header row."""
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def _auto_width(ws):
    """Auto-adjust column widths."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


def export_data(section, data):
    """
    Export data to an Excel file and return bytes.
    section: str — key from SECTION_COLUMNS
    data: list of dicts
    """
    wb = Workbook()
    ws = wb.active
    ws.title = section.replace('_', ' ').title()

    columns = SECTION_COLUMNS.get(section, [])
    if not columns:
        return None

    # Write header
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    _style_header(ws, len(columns))

    # Write data rows
    row_fill_alt = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for row_idx, item in enumerate(data, 2):
        values = _extract_row_values(section, item)
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if row_idx % 2 == 0:
                cell.fill = row_fill_alt

    _auto_width(ws)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _extract_row_values(section, item):
    """Extract row values based on section type."""
    if section == 'invigilators':
        return [
            item.get('name', ''), item.get('email', ''),
            item.get('phone', ''), item.get('department', ''),
            'Yes' if item.get('available', True) else 'No',
            item.get('max_duties', 5), item.get('group_id', 0)
        ]
    elif section == 'rooms':
        return [
            item.get('name', ''), item.get('capacity', 30),
            item.get('floor', ''), item.get('room_type', ''),
            'Yes' if item.get('is_available', True) else 'No',
            item.get('equipment', ''), item.get('buffer_seats', 5)
        ]
    elif section == 'subjects':
        return [
            item.get('name', ''), item.get('code', ''),
            item.get('branch', ''), item.get('student_count', 0),
            item.get('color', '#4A90D9')
        ]
    elif section == 'students':
        return [
            item.get('name', ''), item.get('roll_no', ''),
            item.get('branch', ''), item.get('group_id', 0),
            item.get('email', ''), item.get('phone', '')
        ]
    elif section == 'inventory':
        return [
            item.get('name', ''), item.get('category', ''),
            item.get('quantity', 0), item.get('min_threshold', 10),
            item.get('unit', 'pcs')
        ]
    elif section == 'branches':
        return [
            item.get('name', ''), item.get('code', ''),
            item.get('color', '#4A90D9'), item.get('student_count', 0)
        ]
    elif section == 'staff_duties':
        return [
            item.get('staff_name', ''), item.get('duty_description', ''),
            item.get('location', ''), item.get('date', ''),
            item.get('start_time', ''), item.get('end_time', '')
        ]
    elif section == 'exams':
        return [
            item.get('subject_name', ''), item.get('room_name', ''),
            item.get('date', ''), item.get('start_time', ''),
            item.get('end_time', ''), item.get('session_label', ''),
            item.get('status', '')
        ]
    return []


def import_data(section, file_stream):
    """
    Import data from an Excel file.
    Returns a list of dicts matching the section schema.
    """
    try:
        wb = load_workbook(file_stream, read_only=True, data_only=True)
        ws = wb.active

        columns = SECTION_COLUMNS.get(section, [])
        if not columns:
            return [], 'Unknown section type.'

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if not rows:
            return [], 'No data rows found in the Excel file.'

        data = []
        for row in rows:
            if not row or all(v is None for v in row):
                continue
            item = _parse_row(section, row, columns)
            if item:
                data.append(item)

        wb.close()
        return data, None

    except Exception as e:
        return [], f'Error reading Excel file: {str(e)}'


def _parse_row(section, row, columns):
    """Parse a row of data based on section."""
    def safe_get(idx, default=''):
        try:
            val = row[idx]
            return val if val is not None else default
        except (IndexError, TypeError):
            return default

    def safe_int(val, default=0):
        try:
            if val is None or str(val).strip() == '': return default
            return int(float(str(val).strip()))
        except (ValueError, TypeError):
            return default

    if section == 'invigilators':
        name = str(safe_get(0, '')).strip()
        if not name:
            return None
        return {
            'name': name,
            'email': str(safe_get(1, '')).strip(),
            'phone': str(safe_get(2, '')).strip(),
            'department': str(safe_get(3, '')).strip(),
            'available': str(safe_get(4, 'Yes')).strip().lower() in ('yes', 'true', '1', 'y', '✓'),
            'max_duties': safe_int(safe_get(5), 5),
            'group_id': safe_int(safe_get(6), 0),
        }

    elif section == 'rooms':
        name = str(safe_get(0, '')).strip()
        if not name:
            return None
        return {
            'name': name,
            'capacity': safe_int(safe_get(1), 30),
            'floor': str(safe_get(2, 'Ground')).strip(),
            'room_type': str(safe_get(3, 'Classroom')).strip(),
            'is_available': str(safe_get(4, 'Yes')).strip().lower() in ('yes', 'true', '1', 'y', '✓'),
            'equipment': str(safe_get(5, '')).strip(),
            'buffer_seats': safe_int(safe_get(6), 5),
        }

    elif section == 'subjects':
        name = str(safe_get(0, '')).strip()
        if not name:
            return None
        return {
            'name': name,
            'code': str(safe_get(1, '')).strip(),
            'branch': str(safe_get(2, '')).strip(),
            'student_count': safe_int(safe_get(3), 0),
            'color': str(safe_get(4, '#4A90D9')).strip(),
        }

    elif section == 'students':
        name = str(safe_get(0, '')).strip()
        if not name:
            return None
        return {
            'name': name,
            'roll_no': str(safe_get(1, '')).strip(),
            'branch': str(safe_get(2, '')).strip(),
            'group_id': safe_int(safe_get(3), 0),
            'email': str(safe_get(4, '')).strip(),
            'phone': str(safe_get(5, '')).strip(),
        }

    elif section == 'inventory':
        name = str(safe_get(0, '')).strip()
        if not name:
            return None
        return {
            'name': name,
            'category': str(safe_get(1, 'General')).strip(),
            'quantity': safe_int(safe_get(2), 0),
            'min_threshold': safe_int(safe_get(3), 10),
            'unit': str(safe_get(4, 'pcs')).strip(),
        }

    elif section == 'branches':
        name = str(safe_get(0, '')).strip()
        if not name:
            return None
        return {
            'name': name,
            'code': str(safe_get(1, '')).strip(),
            'color': str(safe_get(2, '#4A90D9')).strip(),
            'student_count': safe_int(safe_get(3), 0),
        }

    elif section == 'staff_duties':
        name = str(safe_get(0, '')).strip()
        if not name:
            return None
        return {
            'staff_name': name,
            'duty_description': str(safe_get(1, '')).strip(),
            'location': str(safe_get(2, '')).strip(),
            'date': str(safe_get(3, '')).strip(),
            'start_time': str(safe_get(4, '09:00')).strip(),
            'end_time': str(safe_get(5, '17:00')).strip(),
        }

    elif section == 'exams':
        subject_name = str(safe_get(0, '')).strip()
        if not subject_name:
            return None
        return {
            'subject_name': subject_name,
            'room_name': str(safe_get(1, '')).strip(),
            'date': str(safe_get(2, '')).strip(),
            'start_time': str(safe_get(3, '09:00')).strip(),
            'end_time': str(safe_get(4, '12:00')).strip(),
            'session_label': str(safe_get(5, 'Morning')).strip(),
            'status': str(safe_get(6, 'scheduled')).strip(),
        }

    elif section == 'attendance':
        # Custom parser for Attendance Sheet Import
        name = str(safe_get(1, '')).strip() # Column 2 in Attendance Sheet
        if not name: return None
        return {
            'name': name,
            'room_or_duty': str(safe_get(2, '')).strip(),
            'date': str(safe_get(3, '')).strip(),
            'attended': str(safe_get(5, '')).strip() == '✓',
            'check_in_time': str(safe_get(6, '')).strip()
        }

    return None


def generate_template(section):
    """Generate a blank Excel template for a section."""
    wb = Workbook()
    ws = wb.active
    ws.title = section.replace('_', ' ').title()

    columns = SECTION_COLUMNS.get(section, [])
    if not columns:
        return None

    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    _style_header(ws, len(columns))

    # Add 3 example rows
    examples = _get_examples(section)
    for row_idx, example in enumerate(examples, 2):
        for col_idx, val in enumerate(example, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    _auto_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _get_examples(section):
    """Return example rows for templates."""
    if section == 'invigilators':
        return [
            ['Dr. Smith', 'smith@college.edu', '9876543210', 'Computer Science', 'Yes', 5, 1],
            ['Prof. Johnson', 'johnson@college.edu', '9876543211', 'Mathematics', 'Yes', 4, 2],
            ['Dr. Williams', 'williams@college.edu', '9876543212', 'Physics', 'Yes', 5, 1],
        ]
    elif section == 'rooms':
        return [
            ['Room 101', 40, '1st Floor', 'Classroom', 'Yes', 'Projector, AC', 5],
            ['Hall A', 120, 'Ground', 'Exam Hall', 'Yes', 'CCTV, AC', 10],
            ['Room 202', 35, '2nd Floor', 'Lab', 'Yes', 'Computers', 3],
        ]
    elif section == 'subjects':
        return [
            ['Data Structures', 'CS301', 'Computer Science', 60, '#4A90D9'],
            ['Linear Algebra', 'MA201', 'Mathematics', 45, '#E74C3C'],
            ['Digital Electronics', 'EC202', 'Electronics', 55, '#2ECC71'],
        ]
    elif section == 'students':
        return [
            ['Alice Johnson', 'CS2024001', 'Computer Science', 1, 'alice@college.edu', '9876543001'],
            ['Bob Smith', 'CS2024002', 'Computer Science', 1, 'bob@college.edu', '9876543002'],
            ['Charlie Brown', 'MA2024001', 'Mathematics', 2, 'charlie@college.edu', '9876543003'],
        ]
    elif section == 'inventory':
        return [
            ['A4 Paper Ream', 'Stationery', 200, 50, 'pcs'],
            ['Blue Pens', 'Stationery', 500, 100, 'pcs'],
            ['Staplers', 'Equipment', 25, 5, 'pcs'],
        ]
    elif section == 'branches':
        return [
            ['Computer Science', 'CS', '#4A90D9', 120],
            ['Electronics', 'EC', '#E74C3C', 90],
            ['Mechanical', 'ME', '#2ECC71', 100],
        ]
    elif section == 'staff_duties':
        return [
            ['Rajesh Kumar', 'Hall Arrangement', 'Exam Hall A', '2026-03-20', '08:00', '17:00'],
            ['Priya Sharma', 'Stationery Distribution', 'Office', '2026-03-20', '08:30', '12:00'],
            ['Amit Patel', 'Student Guidance', 'Main Entrance', '2026-03-20', '08:00', '09:30'],
        ]
    elif section == 'exams':
        return [
            ['Data Structures', 'Hall A', '2026-04-20', '09:00', '12:00', 'Morning', 'scheduled'],
            ['Operating Systems', 'Room 101', '2026-04-20', '14:00', '17:00', 'Afternoon', 'scheduled'],
            ['Digital Electronics', 'Hall B', '2026-04-21', '09:00', '12:00', 'Morning', 'scheduled'],
        ]
    return []


def export_attendance_sheet(duties, title='Duty Attendance'):
    """Generate an attendance Excel sheet with tick-box columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = title

    headers = ['#', 'Name', 'Duty / Room', 'Date', 'Time', 'Attended (✓/✗)', 'Check-in Time', 'Signature']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    _style_header(ws, len(headers))

    for row_idx, duty in enumerate(duties, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=duty.get('name', duty.get('staff_name', duty.get('invigilator_name', ''))))
        ws.cell(row=row_idx, column=3, value=duty.get('room_name', duty.get('duty_description', duty.get('location', ''))))
        ws.cell(row=row_idx, column=4, value=duty.get('date', ''))
        ws.cell(row=row_idx, column=5, value=f"{duty.get('start_time', '')} - {duty.get('end_time', '')}")
        ws.cell(row=row_idx, column=6, value='✓' if duty.get('attended', False) else '✗')
        ws.cell(row=row_idx, column=7, value=duty.get('check_in_time', ''))
        ws.cell(row=row_idx, column=8, value='')

    _auto_width(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
